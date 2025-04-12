import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import win32com.client as win32
from PIL import Image, ImageTk

# -------------------------------
def clean_dataframe(df):
    df.dropna(how='all', inplace=True)
    df.drop_duplicates(inplace=True)
    df.columns = [col.strip().title().replace('_', ' ') for col in df.columns]
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    return df

# -------------------------------
def get_unique_filename(path):
    base, ext = os.path.splitext(path)
    counter = 1
    new_path = path
    while os.path.exists(new_path):
        new_path = f"{base}{counter}{ext}"
        counter += 1
    return new_path

# -------------------------------
def save_excel_with_spacing(df, output_path):
    df.to_excel(output_path, index=False)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    wb.save(output_path)

# -------------------------------
def autofit_excel(file_path):
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(file_path)
        ws = wb.Worksheets(1)
        ws.Columns.AutoFit()
        ws.Rows.AutoFit()
        wb.Save()
        wb.Close()
        excel.Quit()
    except Exception as e:
        print("AutoFit error:", e)

# -------------------------------
def export_pdf(df, output_path):
    c = canvas.Canvas(output_path, pagesize=letter)
    width, height = letter
    x_margin = 40
    y_start = height - 50
    row_height = 20
    font_size = 9
    padding = 10

    c.setFont("Courier", font_size)
    display_df = df.head(25).fillna("")

    col_widths = []
    for col in display_df.columns:
        max_len = max(len(str(col)), *(len(str(cell)) for cell in display_df[col]))
        width_pt = (max_len * 6.5) + padding
        col_widths.append(min(width_pt, 200))

    x = x_margin
    for i, col in enumerate(display_df.columns):
        c.drawString(x, y_start, str(col))
        x += col_widths[i]

    for row_num, row in enumerate(display_df.itertuples(index=False)):
        y = y_start - (row_num + 1) * row_height
        x = x_margin
        for i, cell in enumerate(row):
            c.drawString(x, y, str(cell))
            x += col_widths[i]

    c.save()

# -------------------------------
def process_file(file_path, export_type):
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            df = pd.read_csv(file_path)
        elif ext in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file type.")

        cleaned_df = clean_dataframe(df)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        out_dir = os.path.dirname(file_path)

        if export_type == "excel":
            raw_path = os.path.join(out_dir, f"{base_name}_cleaned.xlsx")
            out_file = get_unique_filename(raw_path)
            save_excel_with_spacing(cleaned_df, out_file)
            autofit_excel(out_file)
        else:
            raw_path = os.path.join(out_dir, f"{base_name}_cleaned.pdf")
            out_file = get_unique_filename(raw_path)
            export_pdf(cleaned_df, out_file)

        print(f"Cleaned: {out_file}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# -------------------------------
def start_gui():
    def select_file():
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])
        if file_paths:
            choice = export_var.get()
            for file_path in file_paths:
                process_file(file_path, choice)

            try:
                import winsound
                winsound.PlaySound("SystemAsterisk", winsound.SND_ALIAS)
            except:
                pass

            messagebox.showinfo("Done", f"Batch cleaned {len(file_paths)} file(s)!")

    root = tk.Tk()
    root.title("Mondo Clean Excel")
    root.geometry("420x360")
    root.resizable(False, False)

    canvas = tk.Canvas(root, width=420, height=360, highlightthickness=0)
    canvas.place(x=0, y=0)

    # Load full-size logo in background
    logo_path = r"D:\VSC\Excel-Auto-Cleaner\assets\MondoCE_faded.png"
    if os.path.exists(logo_path):
        try:
            pil_img = Image.open(logo_path)
            bg_img = ImageTk.PhotoImage(pil_img)
            canvas.bg_img = bg_img  # keep a reference
            canvas.create_image(210, 180, image=bg_img, anchor="center")
        except Exception as e:
            print("Image load error:", e)

    canvas.create_text(210, 40, text="Select files to clean & convert",
                       font=("Impact UI", 18, "bold"), fill="white")

    export_var = tk.StringVar(value="excel")

    rb_excel = tk.Radiobutton(root, text="Export as Excel", variable=export_var, value="excel",
                              font=("Segoe UI", 10))
    rb_pdf = tk.Radiobutton(root, text="Export as PDF", variable=export_var, value="pdf",
                            font=("Segoe UI", 10))

    btn = tk.Button(root, text="Select Files", command=select_file, width=25,
                    font=("Segoe UI", 11))

    canvas.create_window(210, 100, window=rb_excel)
    canvas.create_window(210, 130, window=rb_pdf)
    canvas.create_window(210, 190, window=btn)

    root.mainloop()

# -------------------------------
if __name__ == "__main__":
    start_gui()
